import logging
from pathlib import Path
from typing import List, Dict

import openpyxl

from .models import PhotoMetadata, GPSCoordinates

logger = logging.getLogger(__name__)


class ExcelImporter:
    """Imports photo metadata from an Excel file (Source of Truth).

    Reads headers dynamically (case-insensitive) and supports columns with text
    containing: "Nº", "File", "DESCRIPTION", "Date", "Latitude", "Longitude", "Altitude".
    """

    HEADER_KEYS = {
        # Removed "no" and "n" to avoid matching "Longitud" or "Filename"
        "num": ["nº", "numero", "n°", "id_foto"],
        "archivo": ["archivo", "file", "nombre", "filename"],
        "descripcion": ["descripción", "descripcion", "description", "notas"],
        "fecha": ["fecha", "date", "datetime", "timestamp"],
        "latitud": ["latitud", "lat", "latitude"],
        "longitud": ["longitud", "lon", "long", "lng", "longitude"],
        "altitud": ["altitud", "alt", "altitude", "elevacion"],
        "azimut": ["rumbo", "azimut", "azimuth", "bearing", "direccion"],
    }

    def parse_excel(self, excel_path: Path | str) -> List[PhotoMetadata]:
        excel_path = Path(excel_path)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        # Map headers from row 1
        header_map: Dict[str, int] = {}
        for cell in ws[1]:
            if cell.value is None:
                continue
            text = str(cell.value).strip().lower()
            if not text:
                continue
            for key, variants in self.HEADER_KEYS.items():
                if any(v in text for v in variants):
                    header_map[key] = cell.column  # 1-based index

        logger.info(f"Detected header map: {header_map}")

        # Critical validations
        missing = [k for k in ("archivo", "latitud", "longitud") if k not in header_map]
        if missing:
            raise ValueError(
                f"Missing critical columns in Excel: {', '.join(missing)}. "
                "Make sure to include columns for 'File', 'Latitude', and 'Longitude'."
            )

        results: List[PhotoMetadata] = []

        def _sanitize_cell_value(value):
            """
            Sanitize cell values to prevent CSV/Excel formula injection.
            Only applies to string values that look like formulas.
            Dangerous formulas start with: =, @, tab, carriage return.
            Note: + and - are excluded to allow negative numbers in text form.
            """
            if value is None:
                return None
            if isinstance(value, str):
                value_stripped = value.strip()
                # Only escape truly dangerous formula characters (not + or - which could be numbers)
                dangerous_prefixes = ("=", "@", "\t", "\r")
                if value_stripped.startswith(dangerous_prefixes):
                    logger.warning(f"Sanitized potentially dangerous cell value: {value[:20]}...")
                    return "'" + value  # Escape with single quote
            return value

        for row_idx in range(2, ws.max_row + 1):

            def _val(col_key: str):
                """Get cell value with formula injection sanitization (for text fields)."""
                col = header_map.get(col_key)
                raw_value = ws.cell(row=row_idx, column=col).value if col else None
                return _sanitize_cell_value(raw_value)

            def _val_raw(col_key: str):
                """Get raw cell value without sanitization (for numeric fields)."""
                col = header_map.get(col_key)
                return ws.cell(row=row_idx, column=col).value if col else None

            archivo = _val("archivo")
            if archivo is None:
                continue
            filename = str(archivo).strip()
            if not filename:
                continue

            # Latitud / Longitud obligatorias (use raw values for numeric parsing)
            raw_lat = _val_raw("latitud")
            raw_lon = _val_raw("longitud")

            if raw_lat in (None, "") or raw_lon in (None, ""):
                logger.warning(f"Row {row_idx}: Missing coordinates. Skipping.")
                continue

            try:
                lat = self._to_float(raw_lat)
                lon = self._to_float(raw_lon)
            except ValueError:
                logger.warning(f"Row {row_idx}: Invalid coordinates (Lat/Lon). Skipping.")
                continue

            # Altitud (opcional - use raw for numeric)
            alt_val = _val_raw("altitud")
            try:
                altitude = self._to_float(alt_val) if alt_val not in (None, "") else 0.0
            except ValueError:
                altitude = 0.0

            # Fecha (opcional)
            fecha_cell = _val("fecha")
            timestamp = self._parse_datetime(fecha_cell)

            # Description (optional - sanitize this text field)
            desc_cell = _val("descripcion")
            description = str(desc_cell).strip() if desc_cell not in (None, "") else ""

            # Sequence ID (Nº) - Source of Truth
            raw_num = _val_raw("num")
            sequence_id = str(raw_num).strip() if raw_num is not None else None

            # Azimut (opcional - use raw for numeric)
            az_val = _val_raw("azimut")
            try:
                azimuth = self._to_float(az_val) if az_val not in (None, "") else None
            except ValueError:
                azimuth = None

            coords = GPSCoordinates(latitude=lat, longitude=lon, altitude=altitude, azimuth=azimuth)
            results.append(
                PhotoMetadata(
                    filename=filename,
                    filepath="",  # Will be resolved later from photo index
                    timestamp=timestamp,
                    coordinates=coords,
                    description=description,
                    sequence_id=sequence_id,
                )
            )

        return results

    def _to_float(self, value) -> float:
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(",", ".")
        return float(s)

    def _parse_datetime(self, value):
        # openpyxl usually returns datetime for valid date cells
        from datetime import datetime, date

        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())
        try:
            # Attempt to parse common string formats
            # 2024-01-31 12:34:56 or 31/01/2024 12:34:56
            txt = str(value).strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(txt, fmt)
                except Exception:
                    pass
        except Exception:
            pass
        # If invalid, return None (will be filled later)
        return None
