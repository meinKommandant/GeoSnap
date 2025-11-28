import openpyxl
from openpyxl.styles import Font, Border, Side
import simplekml
from pathlib import Path
from PIL import Image, ImageOps
import pillow_heif
import shutil

# Register HEIF opener
pillow_heif.register_heif_opener()

class ExcelReportGenerator:
    def __init__(self, title="Listado de Fotos"):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self._setup_headers()

    def _setup_headers(self):
        headers = {"B1": "Nº", "C1": "Archivo", "D1": "DESCRIPCIÓN",
                   "E1": "Fecha", "F1": "Latitud", "G1": "Longitud", "H1": "Altitud [m]"}
        
        for cell_coord, text in headers.items():
            cell = self.ws[cell_coord]
            cell.value = text
            cell.font = Font(bold=True)
            cell.border = self.thin_border

        # Anchos de columna
        dims = {'A': 3, 'B': 8, 'C': 30, 'D': 50, 'E': 22, 'F': 15, 'G': 15, 'H': 12}
        for col, width in dims.items():
            self.ws.column_dimensions[col].width = width

    def add_row(self, row_idx, numero_orden, metadata, altitude_val):
        cells = [
            (2, numero_orden), 
            (3, metadata.filename), 
            (4, ""),
            (5, str(metadata.timestamp)), 
            (6, metadata.coordinates.latitude),
            (7, metadata.coordinates.longitude), 
            (8, altitude_val)
        ]
        for col_idx, val in cells:
            c = self.ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = self.thin_border

    def save(self, path):
        self.wb.save(str(path))


class KmzReportGenerator:
    def __init__(self, thumbs_dir: Path):
        self.kml = simplekml.Kml()
        self.thumbs_dir = thumbs_dir
        self._prepare_thumbs_dir()

    def _prepare_thumbs_dir(self):
        if self.thumbs_dir.exists():
            shutil.rmtree(self.thumbs_dir)
        self.thumbs_dir.mkdir(parents=True, exist_ok=True)

    def add_point(self, numero_orden, metadata, img_path, altitude_val):
        # Priority: Excel Sequence ID > Fallback to Loop Counter
        display_id = metadata.sequence_id if metadata.sequence_id else f"{numero_orden} (Auto)"
        titulo_punto = f"Foto Nº {display_id}"
        
        pnt = self.kml.newpoint(name=titulo_punto)
        pnt.coords = [(metadata.coordinates.longitude, metadata.coordinates.latitude)]

        # Estilo del icono (Cámara roja)
        # http://maps.google.com/mapfiles/kml/pal4/icon46.png es una cámara
        pnt.style.iconstyle.icon.href = "http://maps.google.com/mapfiles/kml/pal4/icon46.png"
        pnt.style.iconstyle.color = simplekml.Color.red  # Tinte rojo

        # Miniatura
        thumb_filename = f"thumb_{img_path.name}"
        local_thumb_path = self.thumbs_dir / thumb_filename
        img_src_in_kmz = ""
        
        try:
            with Image.open(img_path) as img:
                img = ImageOps.exif_transpose(img)
                img.thumbnail((800, 800))
                img.save(local_thumb_path, quality=75)
            added_path = self.kml.addfile(str(local_thumb_path))
            if added_path: 
                img_src_in_kmz = added_path.replace("\\", "/")
        except Exception:
            pass

        # HTML Descripción (Tabla con datos)
        # Datos: Nº, Archivo, Descripción, Fecha, Latitud, Longitud, Altitud
        desc_text = (metadata.description or "").strip() if hasattr(metadata, 'description') else ""
        table_html = f"""
        <table border="1" style="border-collapse: collapse; width: 100%;">
            <tr><td><b>Nº</b></td><td>{display_id}</td></tr>
            <tr><td><b>Archivo</b></td><td>{metadata.filename}</td></tr>
            <tr><td><b>DESCRIPCIÓN</b></td><td>{desc_text}</td></tr>
            <tr><td><b>Fecha</b></td><td>{metadata.timestamp}</td></tr>
            <tr><td><b>Latitud</b></td><td>{metadata.coordinates.latitude}</td></tr>
            <tr><td><b>Longitud</b></td><td>{metadata.coordinates.longitude}</td></tr>
            <tr><td><b>Altitud [m]</b></td><td>{altitude_val}</td></tr>
        </table>
        """

        img_html = f'<img src="{img_src_in_kmz}" style="max-width:400px; display:block; margin-bottom:10px;"/>' if img_src_in_kmz else ''
        
        # En el globo mostramos solo la foto y la tabla
        pnt.description = f"{img_html}{table_html}"

    def save(self, path):
        self.kml.savekmz(str(path))
        
    def cleanup(self):
        if self.thumbs_dir.exists():
            shutil.rmtree(self.thumbs_dir)
